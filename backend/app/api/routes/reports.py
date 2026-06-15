import csv
import io
from collections import Counter, defaultdict
from datetime import datetime
from app.core.timezone import frontend_local_to_utc_naive, now_app_tz, utc_naive_to_app
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy import select, or_
from sqlalchemy.orm import Session, selectinload
from app.api.deps import get_current_user
from app.core.config import settings
from app.core.database import get_db
from app.models.ticket import Ticket
from app.models.user import User
from app.services.role_service import has_permission
from app.services.sla_service import refresh_breach, get_sla_snapshot
from app.services.authorization_service import resolve_user_ref

router = APIRouter(prefix='/reports', tags=['reports'])


def _parse_date(value: str | None) -> datetime | None:
    if not value:
        return None
    return frontend_local_to_utc_naive(datetime.fromisoformat(value))


def _ensure_report_access(db: Session, user: User) -> None:
    if not (user.role in ['admin', 'supervisor', 'analyst'] or has_permission(db, user, 'reports', 'view')):
        raise HTTPException(status_code=403, detail='Permiso insuficiente para exportar reportes')
    if not settings.REPORTS_ENABLED:
        raise HTTPException(status_code=403, detail='Reportes deshabilitados por configuración')


def _visible_ticket_query(user: User):
    query = select(Ticket).options(selectinload(Ticket.created_by), selectinload(Ticket.assigned_to))
    if user.role == 'analyst' and user.area:
        query = query.where(or_(Ticket.area_destino == user.area, Ticket.assigned_to_id == user.id))
    return query


def _apply_report_filters(
    query,
    date_from: datetime | None,
    date_to: datetime | None,
    area_destino: str | None,
    project_area: str | None,
    severity: str | None,
    status: str | None,
    assigned_to_id: str | None,
):
    if date_from:
        query = query.where(Ticket.created_at >= date_from)
    if date_to:
        query = query.where(Ticket.created_at <= date_to)
    if area_destino:
        query = query.where(Ticket.area_destino == area_destino)
    if project_area:
        query = query.where(Ticket.project_area == project_area)
    if severity:
        query = query.where(Ticket.severity == severity)
    if status:
        query = query.where(Ticket.status == status)
    if assigned_to_id is not None and assigned_to_id != "":
        if str(assigned_to_id) == "0":
            query = query.where(Ticket.assigned_to_id.is_(None))
        else:
            query = query.where(Ticket.assigned_to_id == int(str(assigned_to_id)))
    return query


def _tickets(
    db: Session,
    user: User,
    date_from: datetime | None,
    date_to: datetime | None,
    area_destino: str | None = None,
    project_area: str | None = None,
    severity: str | None = None,
    status: str | None = None,
    assigned_to_id: str | None = None,
) -> list[Ticket]:
    assigned_filter = assigned_to_id
    if assigned_to_id not in [None, "", "0"] and not str(assigned_to_id).isdigit():
        assigned_user = resolve_user_ref(db, assigned_to_id)
        assigned_filter = str(assigned_user.id) if assigned_user else "-1"
    query = _apply_report_filters(
        _visible_ticket_query(user),
        date_from,
        date_to,
        area_destino,
        project_area,
        severity,
        status,
        assigned_filter,
    )
    rows = list(db.execute(query.order_by(Ticket.created_at.desc())).scalars().unique().all())
    for ticket in rows:
        refresh_breach(ticket)
    db.flush()
    return rows


def _filename(ext: str) -> str:
    return f'ticketera-reporte-avanzado-{now_app_tz().strftime("%Y%m%d-%H%M%S")}.{ext}'


def _local(value: datetime | None) -> str:
    converted = utc_naive_to_app(value)
    return converted.strftime('%Y-%m-%d %H:%M:%S') if converted else ''


def _format_minutes(minutes: float | int | None) -> str:
    if minutes is None:
        return 'N/A'
    minutes = int(round(minutes))
    if minutes < 60:
        return f'{minutes}m'
    hours, mins = divmod(minutes, 60)
    if hours < 24:
        return f'{hours}h {mins}m'
    days, hrs = divmod(hours, 24)
    return f'{days}d {hrs}h {mins}m'


def _summary_from_tickets(db: Session, tickets: list[Ticket]) -> dict:
    total = len(tickets)
    open_tickets = len([t for t in tickets if t.status not in ['Resuelto', 'Cerrado']])
    resolved = len([t for t in tickets if t.status == 'Resuelto'])
    closed = len([t for t in tickets if t.status == 'Cerrado'])
    finished = resolved + closed
    breached = len([t for t in tickets if t.is_sla_breached])

    by_status = Counter(t.status for t in tickets)
    by_severity = Counter(t.severity for t in tickets)
    by_area = Counter(t.area_destino for t in tickets)
    by_requester_area = Counter(t.project_area or 'Sin área' for t in tickets)
    by_category = Counter(t.category or 'Sin categoría' for t in tickets)

    mtta_values: list[float] = []
    mttr_values: list[float] = []
    sla_consumed_values: list[float] = []
    life_values: list[float] = []
    sla_states = Counter()
    workload: dict[str, dict[str, int | str | None]] = defaultdict(lambda: {'analyst': 'Sin asignar', 'analyst_id': None, 'active': 0, 'finished': 0, 'breached': 0})

    for ticket in tickets:
        snap = get_sla_snapshot(db, ticket)
        sla_states[snap.get('sla_state') or 'unknown'] += 1
        if snap.get('sla_consumed_minutes') is not None:
            sla_consumed_values.append(float(snap['sla_consumed_minutes']))
        if snap.get('ticket_age_minutes') is not None:
            life_values.append(float(snap['ticket_age_minutes']))
        if ticket.assigned_at:
            mtta_values.append((ticket.assigned_at - ticket.created_at).total_seconds() / 60)
        end_at = ticket.resolved_at or ticket.closed_at
        if end_at:
            mttr_values.append(max(0, (end_at - ticket.created_at).total_seconds() / 60 - (ticket.sla_paused_seconds or 0) / 60))

        analyst_key = ticket.assigned_to.public_id if ticket.assigned_to else 'unassigned'
        analyst_name = ticket.assigned_to.full_name if ticket.assigned_to else 'Sin asignar'
        row = workload[analyst_key]
        row['analyst'] = analyst_name
        row['analyst_id'] = ticket.assigned_to.public_id if ticket.assigned_to else None
        if ticket.status in ['Resuelto', 'Cerrado']:
            row['finished'] = int(row['finished']) + 1
        else:
            row['active'] = int(row['active']) + 1
        if ticket.is_sla_breached:
            row['breached'] = int(row['breached']) + 1

    sla_compliance = round(((finished - breached) / finished) * 100, 2) if finished else 100.0

    return {
        'kpis': {
            'total_tickets': total,
            'open_tickets': open_tickets,
            'resolved_tickets': resolved,
            'closed_tickets': closed,
            'finished_tickets': finished,
            'breached_tickets': breached,
            'sla_compliance_percent': sla_compliance,
            'mtta_minutes': round(sum(mtta_values) / len(mtta_values), 2) if mtta_values else None,
            'mttr_minutes': round(sum(mttr_values) / len(mttr_values), 2) if mttr_values else None,
            'average_ticket_life_minutes': round(sum(life_values) / len(life_values), 2) if life_values else None,
            'average_sla_consumed_minutes': round(sum(sla_consumed_values) / len(sla_consumed_values), 2) if sla_consumed_values else None,
        },
        'by_status': dict(by_status),
        'by_severity': dict(by_severity),
        'by_area': dict(by_area),
        'sla_states': dict(sla_states),
        'top_requester_areas': [{'area': area, 'count': count} for area, count in by_requester_area.most_common(10)],
        'top_categories': [{'category': category, 'count': count} for category, count in by_category.most_common(10)],
        'analyst_workload': list(workload.values()),
    }


def _build_filtered_report(
    db: Session,
    user: User,
    date_from: str | None,
    date_to: str | None,
    area_destino: str | None,
    project_area: str | None,
    severity: str | None,
    status: str | None,
    assigned_to_id: str | None,
) -> tuple[list[Ticket], dict, dict]:
    start = _parse_date(date_from)
    end = _parse_date(date_to)
    rows = _tickets(db, user, start, end, area_destino, project_area, severity, status, assigned_to_id)
    summary = _summary_from_tickets(db, rows)
    filters = {
        'date_from': date_from,
        'date_to': date_to,
        'area_destino': area_destino,
        'project_area': project_area,
        'severity': severity,
        'status': status,
        'assigned_to_id': assigned_to_id,
    }
    return rows, summary, filters


@router.get('/summary')
def report_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    area_destino: str | None = Query(default=None),
    project_area: str | None = Query(default=None),
    severity: str | None = Query(default=None),
    status: str | None = Query(default=None),
    assigned_to_id: str | None = Query(default=None),
):
    _ensure_report_access(db, current_user)
    tickets, summary, filters = _build_filtered_report(db, current_user, date_from, date_to, area_destino, project_area, severity, status, assigned_to_id)
    return {
        **summary,
        'filters': filters,
        'ticket_sample': [
            {
                'id': ticket.public_id,
                'ticket_number': ticket.ticket_number,
                'subject': ticket.subject,
                'category': ticket.category,
                'severity': ticket.severity,
                'status': ticket.status,
                'area_destino': ticket.area_destino,
                'project_area': ticket.project_area,
                'created_by': ticket.created_by.full_name if ticket.created_by else None,
                'assigned_to': ticket.assigned_to.full_name if ticket.assigned_to else None,
                'created_at': _local(ticket.created_at),
                'resolved_at': _local(ticket.resolved_at or ticket.closed_at),
                'is_sla_breached': ticket.is_sla_breached,
            }
            for ticket in tickets[:20]
        ],
    }


@router.get('/monthly.csv')
def monthly_csv(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    area_destino: str | None = Query(default=None),
    project_area: str | None = Query(default=None),
    severity: str | None = Query(default=None),
    status: str | None = Query(default=None),
    assigned_to_id: str | None = Query(default=None),
):
    _ensure_report_access(db, current_user)
    rows, _, _ = _build_filtered_report(db, current_user, date_from, date_to, area_destino, project_area, severity, status, assigned_to_id)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Ticket', 'Área técnica', 'Área solicitante', 'Severidad', 'Categoría', 'Estado', 'Solicitante', 'Asignado', 'Creación', 'Actualización', 'Resolución/Cierre', 'Vida ticket', 'SLA transcurrido', 'SLA estado', 'SLA vencido'])
    for ticket in rows:
        snap = get_sla_snapshot(db, ticket)
        writer.writerow([
            ticket.public_id,
            ticket.ticket_number,
            ticket.area_destino,
            ticket.project_area,
            ticket.severity,
            ticket.category,
            ticket.status,
            ticket.created_by.email if ticket.created_by else '',
            ticket.assigned_to.email if ticket.assigned_to else '',
            _local(ticket.created_at),
            _local(ticket.updated_at),
            _local(ticket.resolved_at or ticket.closed_at),
            _format_minutes(snap.get('ticket_age_minutes')),
            _format_minutes(snap.get('sla_consumed_minutes')),
            snap.get('sla_state') or '',
            'Sí' if ticket.is_sla_breached else 'No',
        ])
    return Response(content=output.getvalue(), media_type='text/csv; charset=utf-8', headers={'Content-Disposition': f'attachment; filename={_filename("csv")}'})


@router.get('/monthly.xlsx')
def monthly_xlsx(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    area_destino: str | None = Query(default=None),
    project_area: str | None = Query(default=None),
    severity: str | None = Query(default=None),
    status: str | None = Query(default=None),
    assigned_to_id: str | None = Query(default=None),
):
    _ensure_report_access(db, current_user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    tickets, summary, filters = _build_filtered_report(db, current_user, date_from, date_to, area_destino, project_area, severity, status, assigned_to_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen ejecutivo'
    ws.append(['Reporte avanzado de tickets'])
    ws.append(['Generado', now_app_tz().strftime('%Y-%m-%d %H:%M:%S')])
    ws.append([])
    ws.append(['Filtros', 'Valor'])
    for key, value in filters.items():
        ws.append([key, value if value not in [None, ''] else 'Todos'])
    ws.append([])
    ws.append(['Métrica', 'Valor'])
    for key, value in summary['kpis'].items():
        if key.endswith('_minutes') and value is not None:
            value = _format_minutes(value)
        ws.append([key, value])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top')
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14)
    for row in [4, 4 + len(filters) + 2]:
        for cell in ws[row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')

    wd = wb.create_sheet('Distribuciones')
    sections = [
        ('Por estado', summary['by_status']),
        ('Por severidad', summary['by_severity']),
        ('Por área técnica', summary['by_area']),
        ('Estados SLA', summary['sla_states']),
    ]
    for title, values in sections:
        wd.append([title, 'Tickets'])
        wd[wd.max_row][0].font = Font(bold=True)
        wd[wd.max_row][0].fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')
        for key, value in values.items():
            wd.append([key, value])
        wd.append([])
    wd.append(['Top áreas solicitantes', 'Tickets'])
    for item in summary['top_requester_areas']:
        wd.append([item['area'], item['count']])
    wd.append([])
    wd.append(['Top casos de uso', 'Tickets'])
    for item in summary['top_categories']:
        wd.append([item['category'], item['count']])

    wt = wb.create_sheet('Tickets')
    headers = ['ID', 'Ticket', 'Área técnica', 'Área solicitante', 'Severidad', 'Categoría', 'Estado', 'Solicitante', 'Asignado', 'Creación', 'Actualización', 'Resolución/Cierre', 'Vida ticket', 'SLA transcurrido', 'SLA estado', 'SLA vencido']
    wt.append(headers)
    for cell in wt[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')
    for ticket in tickets:
        snap = get_sla_snapshot(db, ticket)
        wt.append([
            ticket.public_id,
            ticket.ticket_number,
            ticket.area_destino,
            ticket.project_area,
            ticket.severity,
            ticket.category,
            ticket.status,
            ticket.created_by.email if ticket.created_by else '',
            ticket.assigned_to.email if ticket.assigned_to else '',
            _local(ticket.created_at),
            _local(ticket.updated_at),
            _local(ticket.resolved_at or ticket.closed_at),
            _format_minutes(snap.get('ticket_age_minutes')),
            _format_minutes(snap.get('sla_consumed_minutes')),
            snap.get('sla_state') or '',
            'Sí' if ticket.is_sla_breached else 'No',
        ])
    for worksheet in [ws, wd, wt]:
        for column_cells in worksheet.columns:
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(42, max(14, max(len(str(cell.value or '')) for cell in column_cells) + 2))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={_filename("xlsx")}'})


@router.get('/monthly.pdf')
def monthly_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    area_destino: str | None = Query(default=None),
    project_area: str | None = Query(default=None),
    severity: str | None = Query(default=None),
    status: str | None = Query(default=None),
    assigned_to_id: str | None = Query(default=None),
):
    _ensure_report_access(db, current_user)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    tickets, summary, filters = _build_filtered_report(db, current_user, date_from, date_to, area_destino, project_area, severity, status, assigned_to_id)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, title='Reporte avanzado ticketera', leftMargin=32, rightMargin=32, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = [Paragraph('Reporte avanzado - Ticketera Ciberseguridad y Networking', styles['Title'])]
    story.append(Paragraph(f'Generado: {now_app_tz().strftime("%Y-%m-%d %H:%M:%S")} · Zona horaria: {settings.APP_TIMEZONE}', styles['Normal']))
    story.append(Spacer(1, 10))

    filter_rows = [['Filtro', 'Valor']] + [[key, str(value if value not in [None, ''] else 'Todos')] for key, value in filters.items()]
    story.append(_styled_table(filter_rows, [150, 320]))
    story.append(Spacer(1, 12))

    kpi_rows = [['Métrica', 'Valor']]
    for key, value in summary['kpis'].items():
        if key.endswith('_minutes') and value is not None:
            value = _format_minutes(value)
        kpi_rows.append([key, str(value)])
    story.append(_styled_table(kpi_rows, [260, 160]))
    story.append(Spacer(1, 14))

    for title, items, label in [
        ('Distribución por estado', summary['by_status'].items(), 'Estado'),
        ('Distribución por severidad', summary['by_severity'].items(), 'Severidad'),
        ('Estados SLA', summary['sla_states'].items(), 'Estado SLA'),
    ]:
        rows = [[label, 'Tickets']] + [[str(key), str(value)] for key, value in items]
        story.append(Paragraph(title, styles['Heading2']))
        story.append(_styled_table(rows, [300, 120]))
        story.append(Spacer(1, 12))

    story.append(PageBreak())
    story.append(Paragraph('Top 10 áreas solicitantes', styles['Heading2']))
    areas = [['Área solicitante', 'Tickets']] + [[item['area'], str(item['count'])] for item in summary['top_requester_areas'][:10]]
    story.append(_styled_table(areas, [320, 100]))
    story.append(Spacer(1, 14))
    story.append(Paragraph('Top 10 casos de uso', styles['Heading2']))
    cats = [['Caso de uso', 'Tickets']] + [[item['category'], str(item['count'])] for item in summary['top_categories'][:10]]
    story.append(_styled_table(cats, [320, 100]))
    story.append(Spacer(1, 14))
    story.append(Paragraph('Muestra de tickets', styles['Heading2']))
    sample = [['Ticket', 'Estado', 'Severidad', 'Área', 'SLA']]
    for ticket in tickets[:20]:
        sample.append([ticket.ticket_number, ticket.status, ticket.severity, ticket.project_area[:30], 'Vencido' if ticket.is_sla_breached else 'OK'])
    story.append(_styled_table(sample, [95, 80, 75, 150, 65]))

    doc.build(story)
    output.seek(0)
    return Response(content=output.getvalue(), media_type='application/pdf', headers={'Content-Disposition': f'attachment; filename={_filename("pdf")}'})


def _styled_table(rows, widths):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    table = Table(rows, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F4C81')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
    ]))
    return table
